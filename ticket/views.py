from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Count
from django.http import JsonResponse, request, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import CreateView, ListView, UpdateView, DeleteView
import openpyxl
from ticket.filters import TicketFilter
from ticket.forms import CreateTicketForm, CreatePrecinctForm, CreateExamForm, FloorCreateForm, RoomCreateForm, \
    CreateExamTypeForm
from ticket.models import Ticket, Floor, Room, ExamType, Precinct, Exam
import logging

logger = logging.getLogger(__name__)


class TicketListView(LoginRequiredMixin, ListView):
    model = Ticket
    template_name = 'ticket/index.html'
    context_object_name = 'tickets'

    def get_queryset(self):
        """Фильтруем билеты по пользователю"""
        queryset = Ticket.objects.filter(user=self.request.user)
        self.filter = TicketFilter(self.request.GET, queryset=queryset, user=self.request.user)  # ✅ Передаем `user`
        return self.filter.qs

    def get_context_data(self, **kwargs):
        """Добавляем фильтр в контекст"""
        context = super().get_context_data(**kwargs)
        context['filter'] = self.filter
        context['ticket_counter'] = Ticket.objects.filter(user=self.request.user).count()
        return context


class CreateTicketView(LoginRequiredMixin, CreateView):
    model = Ticket
    form_class = CreateTicketForm
    template_name = 'ticket/add-ticket.html'
    success_url = reverse_lazy('ticket:index')

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Передаем user
        return kwargs

    def form_valid(self, form):
        """Привязываем билет к пользователю"""
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


class TicketUpdate(LoginRequiredMixin, UpdateView):
    model = Ticket
    form_class = CreateTicketForm
    template_name = 'ticket/update.html'
    success_url = reverse_lazy('ticket:update')

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # ✅ Передаем `user`
        return kwargs

    def get_context_data(self, **kwargs):
        """Добавляем экзамены текущего пользователя и восстанавливаем выбранные значения"""
        context = super().get_context_data(**kwargs)
        ticket = self.object  # Получаем текущий `Ticket`

        # ✅ Фильтруем только экзамены текущего пользователя
        context['form'].fields['exam'].queryset = Exam.objects.filter(user=self.request.user)

        # ✅ Устанавливаем выбранные значения обратно
        context['form'].fields['exam'].initial = ticket.exam
        context['form'].fields['exam_type'].queryset = ExamType.objects.filter(exam=ticket.exam)
        context['form'].fields['exam_type'].initial = ticket.exam_type

        context['form'].fields['precinct'].queryset = Precinct.objects.filter(exam=ticket.exam)
        context['form'].fields['precinct'].initial = ticket.precinct

        context['form'].fields['floor'].queryset = Floor.objects.filter(precinct=ticket.precinct)
        context['form'].fields['floor'].initial = ticket.floor

        context['form'].fields['room'].queryset = Room.objects.filter(floor=ticket.floor, exam_type=ticket.exam_type)
        context['form'].fields['room'].initial = ticket.room

        context['form'].fields['seat_number'].initial = ticket.seat_number

        return context


@login_required
def load_exam_type(request):
    exam = request.GET.get('exam')
    exam_typs = ExamType.objects.filter(exam=exam).values('id', 'name')
    return JsonResponse(list(exam_typs), safe=False)


@login_required
def load_precinct(request):
    exam = request.GET.get('exam')
    precincts = Precinct.objects.filter(exam=exam).values('id', 'name')
    return JsonResponse(list(precincts), safe=False)


@login_required
def load_floor(request):
    precinct_id = request.GET.get('precinct_id')
    floors = Floor.objects.filter(precinct_id=precinct_id).values('id', 'name')
    return JsonResponse(list(floors), safe=False)


@login_required
def load_room(request):
    floor_id = request.GET.get('floor_id')
    exam_type_id = request.GET.get('exam_type')

    # Фильтруем по переданным параметрам
    rooms = Room.objects.all()

    if floor_id:
        rooms = rooms.filter(floor_id=floor_id)
    if exam_type_id:
        rooms = rooms.filter(exam_type_id=exam_type_id)

    # Преобразуем результат в JSON
    rooms = rooms.values('id', 'name')
    return JsonResponse(list(rooms), safe=False)


@login_required
def load_seats(request):
    room_id = request.GET.get('room_id')

    if not room_id:
        return JsonResponse({'error': 'room_id parameter is missing'}, status=400)

    try:
        room = get_object_or_404(Room, id=room_id)
        total_seats = room.capacity
        occupied_seats = Ticket.objects.filter(room=room).values_list('seat_number', flat=True)

        seats = [{"seat_number": i, "occupied": i in occupied_seats} for i in range(1, total_seats + 1)]
        return JsonResponse(seats, safe=False)

    except Exception as e:
        # Добавление отладочного вывода
        print(f"Error loading seats: {e}")
        return JsonResponse({'error': 'Internal Server Error'}, status=500)


@login_required
def get_ticket(request, ticket_id):
    ticket = Ticket.objects.get(id=ticket_id)
    context = {
        'ticket': ticket,
    }
    return render(request, 'ticket/ticket.html', context=context)


@login_required
def delete_ticket(request, ticket_id):
    if request.method == "POST":
        ticket = get_object_or_404(Ticket, id=ticket_id)

        # Увеличиваем количество свободных мест в комнате
        room = ticket.room
        room.available_seats += 1
        room.save()

        # Удаляем билет
        ticket.delete()

        return JsonResponse({"message": "Uğurla silindi"}, status=200)

    return JsonResponse({"error": "Yanlış sorğu metodu"}, status=400)


@login_required
def get_report(request):
    exams = Exam.objects.all()
    context = {
        'exams': exams
    }
    return render(request, 'ticket/report.html', context=context)


@login_required
def load_exam_type_report(request):
    exam = request.GET.get('exam')
    exam_typs = ExamType.objects.filter(exam=exam).values('id', 'name')
    return JsonResponse(list(exam_typs), safe=False)


@login_required
def export_to_excel(request):
    exam_id = request.GET.get('exam')
    exam_type_id = request.GET.get('exam_type')

    if not exam_id or not exam_type_id:
        return HttpResponse("Heç bir imtahan və ya imtahan növü seçilməyib.", status=400)

    tickets = Ticket.objects.filter(exam_id=exam_id, exam_type_id=exam_type_id).select_related('room')

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    if not tickets.exists():

        sheet = workbook.create_sheet(title="Məlumat yoxdur")
        sheet.append(["Seçilmiş imtahan və imtahan növü üçün heç bir məlumat yoxdur."])
    else:

        rooms = tickets.values_list('room', flat=True).distinct()
        for room_id in rooms:
            room = Room.objects.get(id=room_id)
            room_tickets = tickets.filter(room=room)

            sheet = workbook.create_sheet(title=f"OTAQ {room.name}")

            sheet.append(['İş nömrə', 'Ad', 'Soyad', 'Sinif', 'Cins', 'Məktəb', 'Telefon', 'İmtahan növü', 'Yer'])

            for ticket in room_tickets:
                sheet.append([
                    ticket.number,
                    ticket.first_name,
                    ticket.last_name,
                    str(ticket.grader),
                    'Kişi' if ticket.gender == 'M' else 'Qadın',
                    ticket.school,
                    ticket.phone,
                    str(ticket.exam_type),
                    ticket.seat_number,
                ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=İmtahan.xlsx'
    workbook.save(response)

    return response


# EXAM
class ExamListView(LoginRequiredMixin, ListView):
    model = Exam
    template_name = 'ticket/exam/exam-list.html'
    context_object_name = 'exams'

    def get_queryset(self):
        return Exam.objects.filter(user=self.request.user)

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data()
        return context


class ExamCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Exam
    form_class = CreateExamForm
    template_name = 'ticket/exam/add-exam.html'
    success_url = reverse_lazy('ticket:exam_list')
    success_message = 'İmtahan uğurla əlavə edildi!'

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


class ExamUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Exam
    form_class = CreateExamForm  # ✅ Используем нашу форму
    template_name = 'ticket/exam/edit-exam.html'
    success_url = reverse_lazy('ticket:exam_list')
    success_message = 'Məlumatlar uğurla yeniləndi!'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


@login_required
def delete_exam(request, exam_id):
    if request.method == 'POST':
        exam = Exam.objects.get(id=exam_id)

        exam.delete()
        return JsonResponse({"message": "Uğurla silindi"}, status=200)

    return JsonResponse({"error": "Yanlış sorğu metodu"}, status=400)


# EXAM-TYPE
class ExamTypeList(LoginRequiredMixin, ListView):
    model = ExamType
    template_name = 'ticket/exam/exam-type-list.html'
    context_object_name = 'exam_types'

    def get_queryset(self):
        return ExamType.objects.filter(user=self.request.user)

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data()
        return context


class ExamTypeCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = ExamType
    form_class = CreateExamTypeForm
    template_name = 'ticket/exam/add-exam-type.html'
    success_url = reverse_lazy('ticket:exam_type_list')
    success_message = 'İmtahan növü uğurla əlavə edildi!'

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Передаем user
        return kwargs

    def form_valid(self, form):
        """Привязываем Precinct к пользователю"""
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


class ExamTypeUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = ExamType
    form_class = CreateExamTypeForm
    template_name = 'ticket/exam/edit-exam-type.html'
    success_url = reverse_lazy('ticket:exam_type_list')
    success_message = 'Məlumatlar uğurla yeniləndi!'

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Передаем user
        return kwargs

    def form_valid(self, form):
        """Привязываем Precinct к пользователю"""
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


@login_required
def delete_exam_type(request, exam_type_id):
    if request.method == 'POST':
        exam_type = ExamType.objects.get(id=exam_type_id)

        exam_type.delete()
        return JsonResponse({"message": "Uğurla silindi"}, status=200)

    return JsonResponse({"error": "Yanlış sorğu metodu"}, status=400)


# PRECINCT
class PrecinctListView(LoginRequiredMixin, ListView):
    model = Precinct
    template_name = 'ticket/exam/precinct-list.html'
    context_object_name = 'precincts'

    def get_queryset(self):
        """Фильтруем Precinct по текущему пользователю и считаем этажи"""
        return Precinct.objects.filter(user=self.request.user).annotate(floor_count=Count('floor'))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data()
        return context


class CreatePrecinctView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Precinct
    form_class = CreatePrecinctForm
    template_name = 'ticket/exam/add-precinct.html'
    success_url = reverse_lazy('ticket:precinct_list')
    success_message = 'Məntəqə uğurla əlavə edildi!'

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Передаем user
        return kwargs

    def form_valid(self, form):
        """Привязываем Precinct к пользователю"""
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


class PrecinctUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Precinct
    form_class = CreatePrecinctForm
    template_name = 'ticket/exam/edit-precinct.html'
    success_url = reverse_lazy('ticket:precinct_list')
    success_message = 'Məlumatlar uğurla yeniləndi!'

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Передаем user
        return kwargs

    def form_valid(self, form):
        """Привязываем Precinct к пользователю"""
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


@login_required
def delete_precinct(request, precinct_id):
    if request.method == 'POST':
        precinct = Precinct.objects.get(id=precinct_id)

        precinct.delete()
        return JsonResponse({"message": "Uğurla silindi"}, status=200)

    return JsonResponse({"error": "Yanlış sorğu metodu"}, status=400)


# FLOOR
class FloorListView(ListView):
    model = Floor
    template_name = 'ticket/exam/floor-list.html'  # Шаблон для отображения этажей
    context_object_name = 'floors'

    def get_queryset(self):
        """Фильтруем этажи (Floor) только по выбранному Precinct"""
        precinct_id = self.kwargs.get('precinct_id')
        return Floor.objects.filter(precinct_id=precinct_id)

    def get_context_data(self, **kwargs):
        """Добавляем объект Precinct в контекст, чтобы вывести его в шаблоне"""
        context = super().get_context_data(**kwargs)
        context['precinct'] = get_object_or_404(Precinct, id=self.kwargs.get('precinct_id'))
        return context


class FloorCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Floor
    form_class = FloorCreateForm
    template_name = 'ticket/exam/add-floor.html'
    success_url = reverse_lazy('ticket:precinct_list')
    success_message = 'Mərtəbə uğurla əlavə edildi!'

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Передаем user
        return kwargs

    def form_valid(self, form):
        """Привязываем Precinct к пользователю"""
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


class FloorUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Floor
    form_class = FloorCreateForm
    template_name = 'ticket/exam/edit-floor.html'
    success_url = reverse_lazy('ticket:precinct_list')
    success_message = 'Məlumatlar uğurla yeniləndi!'

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Передаем user
        return kwargs

    def form_valid(self, form):
        """Привязываем Precinct к пользователю"""
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


@login_required
def delete_floor(request, floor_id):
    if request.method == 'POST':
        floor = Floor.objects.get(id=floor_id)

        floor.delete()
        return JsonResponse({"message": "Uğurla silindi"}, status=200)

    return JsonResponse({"error": "Yanlış sorğu metodu"}, status=400)


# ROOM
class RoomListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = Room
    template_name = 'ticket/exam/room-list.html'  # Шаблон для списка комнат
    context_object_name = 'rooms'

    def get_queryset(self):
        """Фильтруем комнаты по `floor_id`"""
        floor_id = self.kwargs.get('floor_id')
        return Room.objects.filter(floor_id=floor_id)

    def get_context_data(self, **kwargs):
        """Добавляем `Floor` в контекст для отображения информации"""
        context = super().get_context_data(**kwargs)
        context['floor'] = get_object_or_404(Floor, id=self.kwargs.get('floor_id'))
        return context


class RoomCreateView(CreateView):
    model = Room
    form_class = RoomCreateForm
    template_name = 'ticket/exam/add-room.html'
    success_url = reverse_lazy('ticket:precinct_list')
    success_message = 'Otaq uğurla əlavə edildi!'

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Передаем user
        return kwargs

    def form_valid(self, form):
        """Привязываем Precinct к пользователю"""
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


class RoomUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Room
    form_class = RoomCreateForm
    template_name = 'ticket/exam/edit-room.html'
    success_url = reverse_lazy('ticket:precinct_list')
    success_message = 'Məlumatlar uğurla yeniləndi!'

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Передаем user
        return kwargs

    def form_valid(self, form):
        """Привязываем Precinct к пользователю"""
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        return context


@login_required
def delete_room(request, room_id):
    if request.method == 'POST':
        room = Room.objects.get(id=room_id)

        room.delete()
        return JsonResponse({"message": "Uğurla silindi"}, status=200)

    return JsonResponse({"error": "Yanlış sorğu metodu"}, status=400)
